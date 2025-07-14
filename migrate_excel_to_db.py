import openpyxl
import db_manager
import os

# Caminho para o seu arquivo Excel de estoque
# Certifique-se de que o nome do arquivo e o caminho estão corretos
EXCEL_FILE_PATH = r"C:\Users\Carlao\OneDrive\oneone\ESTOQUE_ABC.xlsx"

def get_numeric_value(cell_value):
    """Converte o valor da célula para int, tratando None ou vazios como 0."""
    if cell_value is None or str(cell_value).strip() == '':
        return 0
    try:
        return int(cell_value)
    except (ValueError, TypeError):
        print(f"Aviso: Valor '{cell_value}' não numérico encontrado e convertido para 0. Verifique a planilha.")
        return 0

def migrate_estoque_from_excel():
    """
    Lê os dados da planilha Excel 'ESTOQUE_ABC.xlsx' de forma estruturada
    e os insere/atualiza no banco de dados PostgreSQL.
    """
    conn = None
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            print(f"Erro: O arquivo Excel '{EXCEL_FILE_PATH}' não foi encontrado.")
            print("Por favor, verifique o caminho e o nome do arquivo.")
            return

        try:
            wb = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
            ws = wb.active
            print(f"Planilha '{EXCEL_FILE_PATH}' lida com sucesso.")
        except PermissionError:
            print(f"Erro de Permissão: Não foi possível acessar o arquivo Excel '{EXCEL_FILE_PATH}'.")
            print("Por favor, certifique-se de que o arquivo não está aberto em outro programa (como o Excel) e que você tem permissão para lê-lo.")
            return
        except Exception as e:
            print(f"Ocorreu um erro ao ler o arquivo Excel: {e}")
            return

        conn = db_manager.connect_db()
        if conn is None:
            print("Não foi possível conectar ao banco de dados. Migração abortada.")
            return

        cursor = conn.cursor()

        def get_model_id(model_name):
            cursor.execute("SELECT modelo_id FROM modelos_cja WHERE nome_modelo = %s;", (model_name,))
            result = cursor.fetchone()
            if result:
                return result[0]
            cursor.execute("INSERT INTO modelos_cja (nome_modelo) VALUES (%s) RETURNING modelo_id;", (model_name,))
            conn.commit()
            print(f"Modelo '{model_name}' inserido na tabela modelos_cja.")
            return cursor.fetchone()[0]

        def get_component_id(component_name, unit='unidade'):
            cursor.execute("SELECT componente_id FROM componentes WHERE nome_componente = %s;", (component_name,))
            result = cursor.fetchone()
            if result:
                return result[0]
            cursor.execute("INSERT INTO componentes (nome_componente, unidade_medida) VALUES (%s, %s) RETURNING componente_id;", (component_name, unit))
            conn.commit()
            print(f"Componente '{component_name}' inserido na tabela componentes.")
            return cursor.fetchone()[0]

        print("Iniciando migração dos dados do estoque...")

        # --- Seção 1: Assento/Encosto Zurich ---
        print("\nProcessando seção: Assento/Encosto Zurich...")
        # Coluna A para Modelo (1)
        # Coluna B para Assento (2) - CORRIGIDO
        # Coluna F para Encosto (6)
        for r_idx in range(7, 11): 
            modelo_nome = ws.cell(row=r_idx, column=1).value # Coluna A
            assento_qty = get_numeric_value(ws.cell(row=r_idx, column=2).value) # Coluna B (Assento) - CORRIGIDO
            encosto_qty = get_numeric_value(ws.cell(row=r_idx, column=6).value) # Coluna F (Encosto)

            if modelo_nome is None:
                print(f"Aviso: Linha {r_idx} (Zurich) ignorada - Modelo CJA não encontrado na Coluna A.")
                continue

            modelo_id = get_model_id(str(modelo_nome).upper())
            
            # Atualiza/Insere Assento Zurich usando 'SET'
            componente_assento_id = get_component_id('Assento Zurich')
            db_manager.insert_or_update_estoque(cursor, conn, modelo_id, componente_assento_id, assento_qty, operation_type='SET')
            print(f"Atualizado/Definido: {modelo_nome} - Assento Zurich (Qtd: {assento_qty})")

            # Atualiza/Insere Encosto Zurich usando 'SET'
            componente_encosto_id = get_component_id('Encosto Zurich')
            db_manager.insert_or_update_estoque(cursor, conn, modelo_id, componente_encosto_id, encosto_qty, operation_type='SET')
            print(f"Atualizado/Definido: {modelo_nome} - Encosto Zurich (Qtd: {encosto_qty})")

        # --- Seção 2: Assento/Encosto Masticmol ---
        print("\nProcessando seção: Assento/Encosto Masticmol...")
        # Coluna A para Modelo (1)
        # Coluna B para Assento (2) - CORRIGIDO
        # Coluna F para Encosto (6)
        for r_idx in range(15, 19): 
            modelo_nome = ws.cell(row=r_idx, column=1).value # Coluna A
            assento_qty = get_numeric_value(ws.cell(row=r_idx, column=2).value) # Coluna B (Assento) - CORRIGIDO
            encosto_qty = get_numeric_value(ws.cell(row=r_idx, column=6).value) # Coluna F (Encosto)

            if modelo_nome is None:
                print(f"Aviso: Linha {r_idx} (Masticmol) ignorada - Modelo CJA não encontrado na Coluna A.")
                continue

            modelo_id = get_model_id(str(modelo_nome).upper())
            
            # Atualiza/Insere Assento Masticmol usando 'SET'
            componente_assento_id = get_component_id('Assento Masticmol')
            db_manager.insert_or_update_estoque(cursor, conn, modelo_id, componente_assento_id, assento_qty, operation_type='SET')
            print(f"Atualizado/Definido: {modelo_nome} - Assento Masticmol (Qtd: {assento_qty})")

            # Atualiza/Insere Encosto Masticmol usando 'SET'
            componente_encosto_id = get_component_id('Encosto Masticmol')
            db_manager.insert_or_update_estoque(cursor, conn, modelo_id, componente_encosto_id, encosto_qty, operation_type='SET')
            print(f"Atualizado/Definido: {modelo_nome} - Encosto Masticmol (Qtd: {encosto_qty})")

        # --- Seção 3: Tampos ---
        print("\nProcessando seção: Tampos...")
        # Coluna A para Modelo (1)
        # Coluna B para Tampo MDF (2)
        # Coluna F para Tampo Plástico (6)
        # Coluna G para Tampo Masticmol (7)
        for r_idx in range(24, 28): 
            modelo_nome = ws.cell(row=r_idx, column=1).value # Coluna A
            tampo_mdf_qty = get_numeric_value(ws.cell(row=r_idx, column=2).value) # Coluna B (MDF)
            tampo_plastico_qty = get_numeric_value(ws.cell(row=r_idx, column=6).value) # Coluna F (PLASTICO)
            tampo_masticmol_qty = get_numeric_value(ws.cell(row=r_idx, column=7).value) # Coluna G (MASTICMOL)

            if modelo_nome is None:
                print(f"Aviso: Linha {r_idx} (Tampos) ignorada - Modelo CJA não encontrado na Coluna A.")
                continue

            modelo_id = get_model_id(str(modelo_nome).upper())
            
            # Atualiza/Insere Tampo MDF usando 'SET'
            componente_tampo_mdf_id = get_component_id('Tampo MDF')
            db_manager.insert_or_update_estoque(cursor, conn, modelo_id, componente_tampo_mdf_id, tampo_mdf_qty, operation_type='SET')
            print(f"Atualizado/Definido: {modelo_nome} - Tampo MDF (Qtd: {tampo_mdf_qty})")

            # Atualiza/Insere Tampo Plastico usando 'SET'
            componente_tampo_plastico_id = get_component_id('Tampo Plastico')
            db_manager.insert_or_update_estoque(cursor, conn, modelo_id, componente_tampo_plastico_id, tampo_plastico_qty, operation_type='SET')
            print(f"Atualizado/Definido: {modelo_nome} - Tampo Plastico (Qtd: {tampo_plastico_qty})")

            # Atualiza/Insere Tampo Masticmol usando 'SET'
            componente_tampo_masticmol_id = get_component_id('Tampo Masticmol')
            db_manager.insert_or_update_estoque(cursor, conn, modelo_id, componente_tampo_masticmol_id, tampo_masticmol_qty, operation_type='SET')
            print(f"Atualizado/Definido: {modelo_nome} - Tampo Masticmol (Qtd: {tampo_masticmol_qty})")

        # --- Seção 4: Porta-Livro (Geral) ---
        print("\nProcessando seção: Porta-Livro (Geral)...")
        # Porta-Livro está na célula K7 (coluna 11)
        porta_livro_qty = get_numeric_value(ws.cell(row=7, column=11).value) # Coluna K
        componente_porta_livro_id = get_component_id('Porta-Livro')
        # Para Porta-Livro, o modelo_id é NULL, pois é um componente geral
        db_manager.insert_or_update_estoque(cursor, conn, None, componente_porta_livro_id, porta_livro_qty, operation_type='SET')
        print(f"Atualizado/Definido: Porta-Livro (Geral) (Qtd: {porta_livro_qty})")

        print("\nMigração do estoque concluída com sucesso!")

    except Exception as e:
        print(f"Ocorreu um erro durante a migração: {e}")
        if conn:
            conn.rollback()
    finally:
        if conn:
            db_manager.close_db_connection(conn)

if __name__ == "__main__":
    migrate_estoque_from_excel()
